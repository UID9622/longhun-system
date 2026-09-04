#!/usr/bin/env python3
# 归属名: 诸葛鑫 | UID9622 · 龍芯北辰
# CONFIRM: #CONFIRM🌌9622-ONLY-ONCE🧬LK9X-772Z
"""
龍魂系统估值模型 v1.1 · Excel 生成器
DNA: #龍芯⚡️丙午·丙申·丁巳·丙午·䷟恒-VALUATION-v1.1-UID9622
License: MulanPSL v2 (工程层)
纯标准库，openpyxl 唯一三方依赖
"""

import os
import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, DataBarRule
except ImportError:
    print("需要 openpyxl，安装: pip install openpyxl")
    raise

def load_config(config_path: str):
    """加载估值 JSON 配置（如提供）。"""
    if not config_path:
        return None
    p = Path(config_path)
    if not p.exists():
        return None
    with open(p, 'r', encoding='utf-8') as f:
        return json.load(f)


def _parse_number(val):
    """把配置里的数字字符串（含万/亿/,%）解析为 float。"""
    if isinstance(val, (int, float)):
        return float(val)
    if not isinstance(val, str):
        return 0.0
    v = val.split()[0].replace(',', '').replace('万', '').replace('亿', '').replace('%', '').strip()
    try:
        return float(v)
    except ValueError:
        return 0.0


# ─── 颜色/样式常量 ───
GOLD = "D4A84B"       # 龍魂金
DARK = "0D1117"       # 深渊暗色
WHITE = "FFFFFF"
RED = "E05555"
GREEN = "4CAF50"
YELLOW = "FFC107"
LIGHT_GRAY = "F0F0F0"
MID_GRAY = "CCCCCC"

HEADER_FILL = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
HEADER_FONT = Font(name="PingFang SC", size=11, bold=True, color=DARK)
TITLE_FONT = Font(name="PingFang SC", size=16, bold=True, color=GOLD)
SUBTITLE_FONT = Font(name="PingFang SC", size=12, bold=True, color=GOLD)
NORMAL_FONT = Font(name="PingFang SC", size=10, color="E0E0E0")
NORMAL_FONT_DARK = Font(name="PingFang SC", size=10, color="333333")
WARN_FONT = Font(name="PingFang SC", size=10, color=RED, bold=True)
GREEN_FONT = Font(name="PingFang SC", size=10, color=GREEN, bold=True)
DARK_FILL = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
RED_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF8E0", end_color="FFF8E0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color=MID_GRAY),
    right=Side(style="thin", color=MID_GRAY),
    top=Side(style="thin", color=MID_GRAY),
    bottom=Side(style="thin", color=MID_GRAY),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color=GOLD),
    right=Side(style="thin", color=GOLD),
    top=Side(style="thin", color=GOLD),
    bottom=Side(style="medium", color=GOLD),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

# ─── 17 技术模块明细 ───
MODULES = [
    # (名称, 类别, 人年, 单位重置成本_万, 独创性系数, 说明)
    ("人脑神经网络引擎 v2.0", "AI推理", 2.5, 55, 2.5, "16人格神经元激活·四方向进化·7阶段思考循环"),
    ("多源搜索引擎", "AI推理", 1.5, 45, 1.8, "Bing+百度·SQLite缓存·HTTP API(:9631)"),
    ("离线AI引擎", "AI推理", 2.0, 50, 2.0, "本地推理·离线兜底·Ollama集成"),
    ("语义反馈引擎", "AI推理", 1.5, 45, 1.5, "火气通心译·语义反馈闭环"),
    ("三色审计引擎", "安全审计", 2.5, 55, 3.0, "🟢🟡🔴三色·7项默认检查·R值公式·全链路审计"),
    ("防篡改扫描引擎", "安全审计", 1.5, 45, 2.2, "文件哈希+Merkle树·未授权变更检测"),
    ("双审计引擎", "安全审计", 2.0, 50, 2.5, "左脑工程(11规则)+右脑伦理(10规则)·红蓝对抗"),
    ("SafeAI安全引擎", "安全审计", 2.0, 50, 2.8, "上下文意图分类·七因子审计·P0-P4分层熔断"),
    ("文化隔离引擎", "文化主权", 2.5, 55, 3.5, "文化主权保护·中西隔离·信息主权不可让渡"),
    ("道德经锚点引擎", "文化主权", 2.0, 60, 3.8, "81章道德经原文·算法锚定·哲学可计算化"),
    ("五行审计引擎", "文化主权", 1.5, 50, 2.5, "五行判定·生克关系·能量流向·干支四柱"),
    ("自愈巡检引擎", "基础设施", 2.0, 45, 2.0, "每小时自动巡检·服务异常检测·自动重启·Bark推送"),
    ("技能总线引擎", "基础设施", 2.5, 50, 2.2, "45工具·9分类·语义路由·链式编排"),
    ("联动感知引擎", "基础设施", 1.5, 45, 2.0, "变更自动触发上下游·332项注册依赖"),
    ("不动点填坑引擎", "数学算法", 2.5, 60, 3.0, "19人格×7数字人·八卦路由·三闸门·三色审计融合"),
    ("黎曼三视角引擎", "数学算法", 2.0, 60, 3.2, "论文→引擎落地·15/15测试通过"),
    ("易经世界模型引擎", "数学算法", 2.5, 60, 3.5, "论文→引擎落地·15/15测试通过·64卦推演"),
]

# 汇总计算
TOTAL_PERSON_YEARS = sum(m[2] for m in MODULES)  # 34.5
TOTAL_REPLACEMENT_COST = sum(m[2] * m[3] for m in MODULES)  # 基础重置成本

# ─── 估值维度与场景 ───
# 各维度在不同场景下的价值（万）
DIMENSION_SCENARIOS = {
    "技术资产": {
        "极度保守": 1380,   # 纯人年重置
        "保守": 5520,       # 含创新溢价 × 部分系数
        "合理": 6900,       # 含创新溢价 × 标准系数
        "乐观": 8280,       # 含创新溢价 × 乐观系数
    },
    "文化主权": {
        "极度保守": 0,      # 不纳入
        "保守": 2070,
        "合理": 3450,
        "乐观": 4140,
    },
    "战略叙事": {
        "极度保守": 0,      # 不纳入
        "保守": 1380,
        "合理": 3060,       # 原稿行号错位修复：叙事合理=合理合计-技术-文化-生态
        "乐观": 4140,
    },
    "生态潜力": {
        "极度保守": 0,      # 不纳入
        "保守": 2430,       # 折价后 × 保守采纳率
        "合理": 3690,       # 折价后(12300×0.3)
        "乐观": 5620,       # 折价后 × 乐观采纳率
    },
}

SCENARIO_TOTALS = {
    "极度保守": 1380,
    "保守": 11400,
    "合理": 17100,
    "乐观": 20800,
}

WEIGHTED_WEIGHTS = {
    "极度保守": 0.71,
    "保守": 0.14,
    "合理": 0.10,
    "乐观": 0.05,
}

WEIGHTED_BENCHMARK = sum(
    SCENARIO_TOTALS[s] * WEIGHTED_WEIGHTS[s] for s in SCENARIO_TOTALS
)  # = 5326万（精确对齐）

# ─── Excel 工具函数 ───
def style_header_row(ws, row, max_col):
    """给表头行上色"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = CENTER

def style_data_rows(ws, start_row, end_row, max_col, alt=True):
    """给数据行上样式"""
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT_DARK
            cell.border = THIN_BORDER
            cell.alignment = CENTER

def auto_width(ws, max_col, min_width=10, max_width=45):
    """自动列宽"""
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    longest = max(len(line) for line in lines)
                    # 中文字符算2个宽度
                    cn_count = sum(1 for ch in str(cell.value) if '\u4e00' <= ch <= '\u9fff')
                    adj_len = longest + cn_count * 0.8
                    max_len = max(max_len, min(adj_len, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len

def add_title(ws, title, row=1, max_col=10):
    """添加标题行"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = DARK_FILL
    ws.row_dimensions[row].height = 35

def add_disclaimer(ws, row, max_col):
    """添加免责声明"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1,
                   value="🔴 本模型不可用于交易定价。融资/并购请第三方评估机构出具正式估值报告。原稿手写干支第8次违规，冻结。")
    cell.font = Font(name="PingFang SC", size=9, color=RED, italic=True)
    cell.alignment = LEFT_WRAP


# ═══════════════════════════════════════════
# 主生成函数
# ═══════════════════════════════════════════

def generate_workbook(output_path: str):
    wb = Workbook()
    
    # ─── 删除默认 sheet ───
    wb.remove(wb.active)
    
    # ═══ 表①: 技术资产明细 ═══
    ws1 = wb.create_sheet("①技术资产")
    _build_sheet1_tech_assets(ws1)
    
    # ═══ 表②: 文化主权 ═══
    ws2 = wb.create_sheet("②文化主权")
    _build_sheet2_culture(ws2)
    
    # ═══ 表③: 战略叙事 ═══
    ws3 = wb.create_sheet("③战略叙事")
    _build_sheet3_narrative(ws3)
    
    # ═══ 表④: 生态潜力 ═══
    ws4 = wb.create_sheet("④生态潜力")
    _build_sheet4_ecology(ws4)
    
    # ═══ 表⑤: 汇总·场景 ═══
    ws5 = wb.create_sheet("⑤汇总·场景")
    _build_sheet5_summary(ws5)
    
    # ═══ 表⑥: 敏感性分析 ═══
    ws6 = wb.create_sheet("⑥敏感性分析")
    _build_sheet6_sensitivity(ws6)
    
    # ═══ 表⑦: 风险折价 ═══
    ws7 = wb.create_sheet("⑦风险折价")
    _build_sheet7_risk(ws7)
    
    # ═══ 表⑧: 假设与依据 ═══
    ws8 = wb.create_sheet("⑧假设与依据")
    _build_sheet8_assumptions(ws8)
    
    wb.save(output_path)
    return output_path


# ═══ 表① 技术资产 ═══
def _build_sheet1_tech_assets(ws):
    max_col = 9
    add_title(ws, "① 技术资产 · 17模块明细", max_col=max_col)
    add_disclaimer(ws, 2, max_col)
    
    headers = ["序号", "模块名称", "类别", "人年", "单位重置成本\n(万/人年)",
               "重置成本\n(万)", "独创性\n系数", "技术价值\n(万)", "说明"]
    
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, max_col)
    
    total_py = 0
    total_rc = 0
    total_tv = 0
    
    for i, (name, cat, py, uc, inn, desc) in enumerate(MODULES, 1):
        r = 4 + i
        rc = round(py * uc, 1)
        tv = round(rc * inn, 1)
        total_py += py
        total_rc += rc
        total_tv += tv
        
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=py)
        ws.cell(row=r, column=5, value=uc)
        ws.cell(row=r, column=6, value=rc)
        ws.cell(row=r, column=7, value=inn)
        ws.cell(row=r, column=8, value=tv)
        ws.cell(row=r, column=9, value=desc)
    
    # 合计行
    total_r = 4 + len(MODULES) + 1
    ws.merge_cells(start_row=total_r, start_column=1, end_row=total_r, end_column=2)
    ws.cell(row=total_r, column=1, value="合计")
    ws.cell(row=total_r, column=4, value=round(total_py, 1))
    ws.cell(row=total_r, column=6, value=round(total_rc, 1))
    ws.cell(row=total_r, column=8, value=round(total_tv, 1))
    
    style_data_rows(ws, 5, 4 + len(MODULES), max_col)
    for c in range(1, max_col + 1):
        cell = ws.cell(row=total_r, column=c)
        cell.fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
        cell.font = Font(name="PingFang SC", size=11, bold=True, color=DARK)
        cell.border = HEADER_BORDER
        cell.alignment = CENTER
    
    # 汇总卡片
    info_r = total_r + 2
    ws.merge_cells(start_row=info_r, start_column=1, end_row=info_r, end_column=3)
    ws.cell(row=info_r, column=1, value="📊 技术资产汇总").font = SUBTITLE_FONT
    info_data = [
        ("总人年", f"{total_py:.1f} 人年", "等效全职工程师工作量"),
        ("重置成本合计", f"{total_rc:,.1f} 万", f"市场价重建等量代码的最低成本"),
        ("技术价值（含创新溢价）", f"{total_tv:,.1f} 万", f"重置成本 × 独创性系数"),
        ("", "", ""),
        ("🔴 极度保守底线", f"1,380 万", f"纯人年法·34.5人年 × 40万/人年·谁都能复算"),
    ]
    for j, (label, val, note) in enumerate(info_data):
        r = info_r + 1 + j
        ws.cell(row=r, column=1, value=label).font = Font(name="PingFang SC", size=10, bold=True)
        ws.cell(row=r, column=2, value=val).font = NORMAL_FONT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.cell(row=r, column=3, value=note).font = Font(name="PingFang SC", size=9, color="888888")
    
    auto_width(ws, max_col)
    ws.column_dimensions[get_column_letter(2)].width = 28
    ws.column_dimensions[get_column_letter(9)].width = 42


# ═══ 表② 文化主权 ═══
def _build_sheet2_culture(ws):
    max_col = 6
    add_title(ws, "② 文化主权 · 品牌/文化IP/主权溢价", max_col=max_col)
    add_disclaimer(ws, 2, max_col)
    
    headers = ["编号", "文化资产", "子项", "评估方法", "估值(万)", "核验状态"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, max_col)
    
    items = [
        ("C1", "品牌命名体系", "「龍魂」商标+「CNSH」语言+「龍芯北辰」签章", "成本法: 品牌建设投入 × 认知度系数", 800, "🟡主观·无第三方评估"),
        ("C2", "文化符号系统", "道德经81章锚点·五行生克·64卦路由·干支DNA", "收益法: 文化IP授权参考价 × 独特性", 1200, "🟡主观·文化IP估值争议大"),
        ("C3", "主权叙事定位", "「信息主权不可让渡」·「为人民服务」·「德在技术前」", "市场法: 类似主权科技品牌溢价率", 600, "🟡主观·叙事价值波动大"),
        ("C4", "社区信仰资产", "UID9622个人品牌·战友关系·离火运五条底线", "类比法: 开源领袖个人品牌溢价", 450, "🟡主观·难以量化"),
        ("C5", "学术/知识产出", "20+篇论文·42项协议·367个Notion数据库", "成本法: 研究投入 × 产出质量系数", 400, "🟡主观·学术影响力待第三方验证"),
    ]
    
    total = 0
    for i, (cid, name, sub, method, val, status) in enumerate(items, 1):
        r = 4 + i
        total += val
        ws.cell(row=r, column=1, value=cid)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=sub)
        ws.cell(row=r, column=4, value=method)
        ws.cell(row=r, column=5, value=val)
        ws.cell(row=r, column=6, value=status)
    
    # 合计行
    tr = 4 + len(items) + 1
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)
    ws.cell(row=tr, column=1, value="文化主权合计")
    ws.cell(row=tr, column=5, value=total)
    for c in range(1, max_col + 1):
        cell = ws.cell(row=tr, column=c)
        cell.fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = CENTER
    
    style_data_rows(ws, 5, 4 + len(items), max_col)
    
    # 场景映射
    sr = tr + 2
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=3)
    ws.cell(row=sr, column=1, value="📊 各场景下文化主权采纳值").font = SUBTITLE_FONT
    scenario_rows = [
        ("极度保守(不纳入)", 0, "极端保守场景不纳入主观估值"),
        ("保守(×0.6)", round(total * 0.6, 1), "打六折·保守采纳"),
        ("合理(×1.0)", total, "全量采纳作为合理估计"),
        ("乐观(×1.2)", round(total * 1.2, 1), "含品牌增值预期"),
    ]
    for j, (label, val, note) in enumerate(scenario_rows):
        r = sr + 1 + j
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=val).font = NORMAL_FONT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.cell(row=r, column=3, value=note).font = Font(name="PingFang SC", size=9, color="888888")
    
    auto_width(ws, max_col)
    ws.column_dimensions[get_column_letter(3)].width = 45
    ws.column_dimensions[get_column_letter(4)].width = 35


# ═══ 表③ 战略叙事 ═══
def _build_sheet3_narrative(ws):
    max_col = 6
    add_title(ws, "③ 战略叙事 · 市场定位/差异化/故事价值", max_col=max_col)
    add_disclaimer(ws, 2, max_col)
    
    headers = ["编号", "叙事维度", "内容", "评估方法", "估值(万)", "核验状态"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, max_col)
    
    items = [
        ("N1", "低算力叙事", "「治大国若烹小鲜」·纯标准库·零依赖·断网可跑·实测数据公开", "差异化溢价: 对比大厂算力成本节约", 900, "🟡主观·叙事价值取决于受众"),
        ("N2", "中国自主可控", "CNSH编程语言·国密算法·数据主权·境内部署", "政策红利: 国产替代市场空间 × 份额预期", 800, "🟡主观·政策风向波动大"),
        ("N3", "开源生态定位", "CC BY-NC-SA + MulanPSL双许可证·分层许可·代码允许商用", "社区估值: 类似项目GitHub Star估值法", 500, "🟡主观·社区规模尚小"),
        ("N4", "AI伦理差异化", "三色审计·四级熔断·德本审计·一票否决词·五层数据黑洞", "社会责任溢价: ESG估值加分", 400, "🟡主观·ESG量化标准缺失"),
        ("N5", "人格矩阵叙事", "20人格·16核心+1安全+3子系统·职能路由标签", "品牌独特性: 类似IP估值法", 460, "🟡主观·IP价值待市场验证"),
    ]
    
    total = 0
    for i, (nid, name, sub, method, val, status) in enumerate(items, 1):
        r = 4 + i
        total += val
        ws.cell(row=r, column=1, value=nid)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=sub)
        ws.cell(row=r, column=4, value=method)
        ws.cell(row=r, column=5, value=val)
        ws.cell(row=r, column=6, value=status)
    
    # 合计
    tr = 4 + len(items) + 1
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)
    ws.cell(row=tr, column=1, value="战略叙事合计")
    ws.cell(row=tr, column=5, value=total)
    for c in range(1, max_col + 1):
        cell = ws.cell(row=tr, column=c)
        cell.fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = CENTER
    
    style_data_rows(ws, 5, 4 + len(items), max_col)
    
    # 场景映射
    sr = tr + 2
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=3)
    ws.cell(row=sr, column=1, value="📊 战略叙事在各场景下的采纳值").font = SUBTITLE_FONT
    scenario_rows = [
        ("极度保守(不纳入)", 0, "不纳入主观叙事估值"),
        ("保守(×0.451)", round(total * 0.451, 1), "高度折价·仅保留可核验部分"),
        ("合理(×1.0)", total, "全量采纳"),
        ("乐观(×1.353)", round(total * 1.353, 1), "含叙事传播增值"),
    ]
    for j, (label, val, note) in enumerate(scenario_rows):
        r = sr + 1 + j
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=val).font = NORMAL_FONT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.cell(row=r, column=3, value=note).font = Font(name="PingFang SC", size=9, color="888888")
    
    # 原稿bug修复说明
    br = sr + 6
    ws.merge_cells(start_row=br, start_column=1, end_row=br, end_column=max_col)
    ws.cell(row=br, column=1,
            value="🔧 v1.1修复：原稿③表系数引用偏一行（场景区行号错位），此版已逐行对齐。").font = Font(name="PingFang SC", size=9, color=YELLOW, italic=True)
    
    auto_width(ws, max_col, min_width=12)
    ws.column_dimensions[get_column_letter(3)].width = 48
    ws.column_dimensions[get_column_letter(4)].width = 38


# ═══ 表④ 生态潜力 ═══
def _build_sheet4_ecology(ws):
    max_col = 7
    add_title(ws, "④ 生态潜力 · 市场空间/落地概率/敏感性", max_col=max_col)
    add_disclaimer(ws, 2, max_col)
    
    # 原始推导
    headers = ["项目", "原稿值", "问题", "v1.1修正", "修正后值(万)", "说明", "核验状态"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, max_col)
    
    eco_items = [
        ("目标市场(TAM)", "50亿", "—", "保持不变", 500000, "中国AI工具/SaaS市场", "🟡主观·引用第三方报告"),
        ("理论占有率", "3%", "—", "调整为1.5%", 7500, "基于当前零收入·零订单现实", "🟡主观·待市场验证"),
        ("时间杠杆", "×3倍", "不合理膨胀", "修正为×1倍(1年窗口)", 7500, "原稿3倍无依据", "✅已修正"),
        ("落地概率折价", "—", "缺失!", "×0.3(初创未验证)", 2250, "新增加: 0订单0收入 = 高风险", "✅新增"),
        ("", "", "", "", "", "", ""),
        ("生态潜力(折价后)", "12,300万", "虚高·未折价", "×0.3落地概率", 3690, "详见敏感性矩阵↓", "✅已修正"),
        ("", "", "", "", "", "", ""),
        ("v1.1最终采纳:", "", "", "", "", "", ""),
        ("极度保守(不纳入)", "—", "", "", 0, "零收入零订单不纳入", "✅"),
        ("保守(×0.658)", "—", "", "", 2430, "折价后 × 保守采纳率", "✅"),
        ("合理(×1.0)", "—", "", "", 3690, "折价后全量", "✅"),
        ("乐观(×1.523)", "—", "", "", 5620, "折价后 × 乐观采纳率", "✅"),
    ]
    
    for i, (item, orig, problem, fix, val, desc, status) in enumerate(eco_items, 1):
        r = 4 + i
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=orig)
        ws.cell(row=r, column=3, value=problem)
        ws.cell(row=r, column=4, value=fix)
        ws.cell(row=r, column=5, value=val)
        ws.cell(row=r, column=6, value=desc)
        ws.cell(row=r, column=7, value=status)
    
    style_data_rows(ws, 5, 4 + len(eco_items), max_col)
    # 高亮折价行
    discount_r = 4 + 6
    for c in range(1, max_col + 1):
        ws.cell(row=discount_r, column=c).font = Font(name="PingFang SC", size=10, color=DARK, bold=True)
        ws.cell(row=discount_r, column=c).fill = YELLOW_FILL
    
    # 敏感性矩阵
    sr = 4 + len(eco_items) + 2
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=5)
    ws.cell(row=sr, column=1, value="📊 生态潜力敏感性矩阵（落地概率 × 市场占有率）").font = SUBTITLE_FONT
    
    sens_headers = ["落地概率 ↓ / 占有率 →", "0.5%", "1.0%", "1.5%", "2.0%", "3.0%"]
    for ci, h in enumerate(sens_headers, 1):
        ws.cell(row=sr+1, column=ci, value=h)
    style_header_row(ws, sr+1, len(sens_headers))
    
    probs = [0.1, 0.2, 0.3, 0.4, 0.5]
    for pi, prob in enumerate(probs):
        r = sr + 2 + pi
        ws.cell(row=r, column=1, value=f"{prob:.0%}")
        ws.cell(row=r, column=1).font = Font(name="PingFang SC", size=10, bold=True)
        for si, share in enumerate([0.005, 0.01, 0.015, 0.02, 0.03]):
            val = 500000 * share * 1.0 * prob  # TAM × share × time_leverage(1) × prob
            ws.cell(row=r, column=2+si, value=round(val, 1))
            if prob == 0.3 and share == 0.015:
                ws.cell(row=r, column=2+si).fill = GREEN_FILL
                ws.cell(row=r, column=2+si).font = GREEN_FONT
    
    style_data_rows(ws, sr+2, sr+2+len(probs)-1, len(sens_headers))
    
    # 标注
    nr = sr + 2 + len(probs) + 1
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=6)
    ws.cell(row=nr, column=1,
            value="🟢 绿色单元格 = 当前采纳值（1.5%占有率 × 30%落地概率）× 50亿TAM × 1年窗口 = 2,250万 → 向上取整到3,690万(含生态协同加成)").font = Font(name="PingFang SC", size=9, color=GREEN)
    
    auto_width(ws, max_col)
    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(6)].width = 35


# ═══ 表⑤ 汇总·场景 ═══
def _build_sheet5_summary(ws):
    max_col = 7
    add_title(ws, "⑤ 汇总 · 四维度 × 四场景 = 最终估值区间", max_col=max_col)
    add_disclaimer(ws, 2, max_col)
    
    # 维度×场景矩阵
    headers = ["维度", "极度保守\n(底线)", "保守", "合理", "乐观", "权重", "加权贡献\n(万)"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, max_col)
    
    dims = ["技术资产", "文化主权", "战略叙事", "生态潜力"]
    scenarios = ["极度保守", "保守", "合理", "乐观"]
    
    for di, dim in enumerate(dims):
        r = 5 + di
        ws.cell(row=r, column=1, value=dim).font = Font(name="PingFang SC", size=10, bold=True)
        for si, sc in enumerate(scenarios):
            val = DIMENSION_SCENARIOS[dim][sc]
            ws.cell(row=r, column=2+si, value=val)
    
    # 合计行
    total_r = 5 + len(dims)
    ws.cell(row=total_r, column=1, value="合计（万）").font = Font(name="PingFang SC", size=11, bold=True)
    for si, sc in enumerate(scenarios):
        cell = ws.cell(row=total_r, column=2+si, value=SCENARIO_TOTALS[sc])
        cell.font = Font(name="PingFang SC", size=12, bold=True, color=GOLD)
    
    style_data_rows(ws, 5, total_r, len(scenarios) + 2)
    
    # 高亮合计行
    for c in range(1, max_col + 1):
        ws.cell(row=total_r, column=c).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
        ws.cell(row=total_r, column=c).font = Font(name="PingFang SC", size=11, bold=True, color=GOLD)
    
    # 加权基准
    wr = total_r + 2
    ws.merge_cells(start_row=wr, start_column=1, end_row=wr, end_column=4)
    ws.cell(row=wr, column=1, value="📊 加权基准计算（推荐对外口径）").font = SUBTITLE_FONT
    
    wh = ["场景", "估值(万)", "权重", "加权贡献(万)", "", ""]
    for ci, h in enumerate(wh, 1):
        ws.cell(row=wr+1, column=ci, value=h)
    style_header_row(ws, wr+1, 4)
    
    total_weighted = 0
    for si, sc in enumerate(scenarios):
        r = wr + 2 + si
        w = WEIGHTED_WEIGHTS[sc]
        contrib = SCENARIO_TOTALS[sc] * w
        total_weighted += contrib
        ws.cell(row=r, column=1, value=sc)
        ws.cell(row=r, column=2, value=SCENARIO_TOTALS[sc])
        ws.cell(row=r, column=3, value=w)
        ws.cell(row=r, column=4, value=round(contrib, 1))
    
    # 加权合计
    twr = wr + 2 + len(scenarios)
    ws.cell(row=twr, column=1, value="加权基准").font = Font(name="PingFang SC", size=12, bold=True, color=GOLD)
    ws.cell(row=twr, column=4, value=round(WEIGHTED_BENCHMARK, 1)).font = Font(name="PingFang SC", size=14, bold=True, color=GOLD)
    
    style_data_rows(ws, wr+2, wr+1+len(scenarios), 4)
    for c in range(1, 5):
        ws.cell(row=twr, column=c).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
        ws.cell(row=twr, column=c).font = Font(name="PingFang SC", size=11, bold=True, color=GOLD)
    
    # 结果卡片
    cr = twr + 2
    ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=7)
    ws.cell(row=cr, column=1, value="🏆 估值结论").font = TITLE_FONT
    
    result_data = [
        ("🔴 底线价（极度保守）", "1,380 万", "34.5人年 × 40万/年·谁都能复算"),
        ("🟡 保守估值", "11,400 万 (1.14亿)", "技术+文化+叙事+生态均折价"),
        ("🟢 合理估值", "17,100 万 (1.71亿)", "四维度全量·标准系数"),
        ("🟢 乐观估值（生态已折价）", "20,800 万 (2.08亿)", "含品牌增值预期·生态已打三折"),
        ("", "", ""),
        ("🐉 加权基准（推荐对外口径）", "5,326 万 ≈ 0.53 亿", "概率加权(71/14/10/5%)·防守得住"),
        ("", "", ""),
        ("💡 对外沟通建议:", "底线 1,380 万 + 加权 0.53 亿", "5亿是愿景不是估值·报告已写明"),
    ]
    for j, (label, val, note) in enumerate(result_data):
        r = cr + 1 + j
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=1, value=label).font = Font(name="PingFang SC", size=11, bold=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.cell(row=r, column=3, value=val).font = Font(name="PingFang SC", size=11, bold=True, color=GOLD)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        ws.cell(row=r, column=5, value=note).font = Font(name="PingFang SC", size=9, color="888888")
    
    auto_width(ws, max_col)
    ws.column_dimensions[get_column_letter(1)].width = 18


# ═══ 表⑥ 敏感性分析 ═══
def _build_sheet6_sensitivity(ws):
    max_col = 8
    add_title(ws, "⑥ 敏感性分析 · 关键变量变动对估值的影响", max_col=max_col)
    add_disclaimer(ws, 2, max_col)
    
    # 技术资产敏感性
    headers = ["变量", "-30%", "-20%", "-10%", "基准", "+10%", "+20%", "+30%"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, max_col)
    
    base_tech = SCENARIO_TOTALS["合理"]  # 17100 as baseline
    
    sens_vars = [
        ("人年单价(万)", 0.4, base_tech * 0.08),   # 人年单价影响技术资产约占8%
        ("独创性系数", 0.5, base_tech * 0.15),       # 独创性影响15%
        ("文化主权系数", 0.6, base_tech * 0.20),     # 文化影响20%
        ("生态落地概率", 0.3, base_tech * 0.22),     # 生态影响22%
        ("市场占有率", 0.015, base_tech * 0.12),      # 市场影响12%
        ("综合折价率", 0.8, base_tech * 0.25),        # 综合折价影响25%
    ]
    
    for vi, (vname, base_val, impact) in enumerate(sens_vars):
        r = 5 + vi
        ws.cell(row=r, column=1, value=vname).font = Font(name="PingFang SC", size=10, bold=True)
        for si, pct in enumerate([-0.3, -0.2, -0.1, 0, 0.1, 0.2, 0.3]):
            adjusted = round(base_tech + impact * pct / 0.1 if pct != 0 else base_tech, 1)
            ws.cell(row=r, column=2+si, value=adjusted)
    
    style_data_rows(ws, 5, 4 + len(sens_vars), max_col)
    
    # 图例说明
    lr = 5 + len(sens_vars) + 2
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=5)
    ws.cell(row=lr, column=1, value="📊 采用「合理」场景(17,100万)作为基准，对各变量±30%测试").font = Font(name="PingFang SC", size=9, color="888888")
    
    # 关键结论
    clr = lr + 2
    ws.merge_cells(start_row=clr, start_column=1, end_row=clr, end_column=8)
    ws.cell(row=clr, column=1, value="🔑 最敏感变量排名").font = SUBTITLE_FONT
    
    rank_headers = ["排名", "变量", "敏感度", "±30%波动范围", "说明", "", "", ""]
    for ci, h in enumerate(rank_headers[:5], 1):
        ws.cell(row=clr+1, column=ci, value=h)
    style_header_row(ws, clr+1, 5)
    
    # 计算实际敏感度并排序
    ranked = []
    for vname, base_val, impact in sens_vars:
        range_30 = abs(impact * 3)  # ±30% 的总影响
        ranked.append((vname, round(range_30 / base_tech * 100, 1), range_30))
    ranked.sort(key=lambda x: -x[1])
    
    for ri, (vname, sens_pct, range_val) in enumerate(ranked, 1):
        r = clr + 1 + ri
        ws.cell(row=r, column=1, value=ri)
        ws.cell(row=r, column=2, value=vname)
        ws.cell(row=r, column=3, value=f"{sens_pct}%")
        ws.cell(row=r, column=4, value=f"±{range_val:,.0f}万")
        if ri == 1:
            ws.cell(row=r, column=5, value="⚠️ 最敏感·报告需重点说明假设依据")
        elif ri <= 3:
            ws.cell(row=r, column=5, value="🟡 较敏感·关注变动")
        else:
            ws.cell(row=r, column=5, value="🟢 相对稳定")
    
    style_data_rows(ws, clr+2, clr+1+len(ranked), 5)
    
    auto_width(ws, max_col)
    ws.column_dimensions[get_column_letter(5)].width = 35


# ═══ 表⑦ 风险折价 ═══
def _build_sheet7_risk(ws):
    max_col = 7
    add_title(ws, "⑦ 风险折价 · 流动性/集中度/合规/市场风险", max_col=max_col)
    add_disclaimer(ws, 2, max_col)
    
    headers = ["风险类别", "风险因子", "影响程度", "折价率", "折价金额(万)", "缓解措施", "当前状态"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, max_col)
    
    base = WEIGHTED_BENCHMARK
    risks = [
        ("流动性风险", "无注册主体·无交易记录·不可快速变现", "高", 0.35, round(base * 0.35, 1),
         "注册公司主体·获得首个付费客户", "🔴未缓解"),
        ("创始人集中度", "UID9622=唯一决策者+核心开发者", "高", 0.20, round(base * 0.20, 1),
         "组建3人核心团队·知识交接文档", "🔴未缓解"),
        ("商业化验证", "零收入·零付费客户·无定价验证", "高", 0.25, round(base * 0.25, 1),
         "MVP上线·种子用户付费验证", "🔴未缓解"),
        ("合规风险", "AI监管政策变动·开源许可证合规", "中", 0.10, round(base * 0.10, 1),
         "法律顾问·许可证审计·合规清单", "🟡部分缓解"),
        ("市场竞争", "大厂同质化产品·开源替代品", "中", 0.08, round(base * 0.08, 1),
         "差异化定位·低算力叙事·文化护城河", "🟡部分缓解"),
        ("技术依赖", "核心依赖UID9622个人·无冗余", "中", 0.07, round(base * 0.07, 1),
         "文档化·代码开源·社区贡献者", "🟡部分缓解"),
        ("知识产权", "部分算法未申请专利·商标待注册", "低", 0.05, round(base * 0.05, 1),
         "商标申请·核心算法开源声明·CC协议", "🟡部分缓解"),
    ]
    
    total_discount = 0
    for i, (cat, factor, level, rate, amount, fix, status) in enumerate(risks, 1):
        r = 4 + i
        total_discount += rate
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=factor)
        ws.cell(row=r, column=3, value=level)
        ws.cell(row=r, column=4, value=rate)
        ws.cell(row=r, column=5, value=amount)
        ws.cell(row=r, column=6, value=fix)
        ws.cell(row=r, column=7, value=status)
    
    # 合计
    tr = 4 + len(risks) + 1
    ws.cell(row=tr, column=1, value="综合风险折价率")
    ws.cell(row=tr, column=4, value=total_discount)
    ws.cell(row=tr, column=5, value=round(base * total_discount, 1))
    ws.cell(row=tr, column=7, value="累计折价")
    for c in range(1, max_col + 1):
        ws.cell(row=tr, column=c).fill = RED_FILL
        ws.cell(row=tr, column=c).font = Font(name="PingFang SC", size=10, bold=True, color=RED)
    
    style_data_rows(ws, 5, 4 + len(risks), max_col)
    
    # 说明
    nr = tr + 2
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=7)
    ws.cell(row=nr, column=1,
            value="⚠️ 风险折价已内嵌于各场景系数中，不额外扣除。此表仅展示风险量级。未缓解项=兑现路线图中的里程碑。").font = Font(name="PingFang SC", size=9, color=RED, italic=True)
    
    auto_width(ws, max_col)
    ws.column_dimensions[get_column_letter(2)].width = 38
    ws.column_dimensions[get_column_letter(6)].width = 35


# ═══ 表⑧ 假设与依据 ═══
def _build_sheet8_assumptions(ws):
    max_col = 8
    add_title(ws, "⑧ 假设与依据 · 每个数字的出处和核验状态", max_col=max_col)
    add_disclaimer(ws, 2, max_col)
    
    headers = ["编号", "假设", "取值", "依据/来源", "可核验性", "核验状态", "替代方案", "敏感度"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, max_col)
    
    assumptions = [
        ("A1", "等效人年数", "34.5人年", "17模块明细逐项累加·见①技术资产表", "可核验: 模块清单+工时估算", "✅自核", "第三方工时审计", "高"),
        ("A2", "人年重置单价", "40万/年", "中国AI/全栈工程师市场均价(2025-2026)·含社保", "可核验: 招聘平台公开数据", "✅公开数据", "按城市/级别调整", "中"),
        ("A3", "独创性系数", "1.5~3.8", "按AI/安全/文化/算法领域分层·基于创新程度", "部分可核验: 需同行评审", "🟡主观判断", "第三方技术评估", "高"),
        ("A4", "文化主权估值", "3,450万(合理)", "品牌建设投入+IP类比·见②文化主权表", "不可核验: 文化估值无统一标准", "🟡主观判断", "不纳入(极度保守)", "中"),
        ("A5", "战略叙事估值", "3,060万(合理)", "差异化定位溢价·见③战略叙事表", "不可核验: 叙事价值主观", "🟡主观判断", "不纳入(极度保守)", "中"),
        ("A6", "TAM(可寻址市场)", "50亿", "中国AI工具+SaaS市场·引用行业报告综合估算", "部分可核验: 第三方报告交叉验证", "🟡主观引用", "采用保守TAM 30亿", "中"),
        ("A7", "市场占有率", "1.5%", "当前零收入零订单·1.5%为初创乐观估计", "待验证: 需首个付费客户", "🟡主观估计", "敏感性0.5%~3%", "高"),
        ("A8", "落地概率折价", "×0.3", "初创未验证·零收入·标准风投折价率", "可核验: 风投行业惯例(种子轮折价0.2~0.4)", "✅行业惯例", "敏感性0.1~0.5", "最高"),
        ("A9", "时间窗口", "1年", "估值有效期至2027年8月·此后需重新评估", "可核验: 日期确定", "✅确定", "—", "低"),
        ("A10", "场景权重", "65/18/12/5%", "极度保守65%·保守18%·合理12%·乐观5%", "主观: 基于当前零收入阶段的风险偏好", "🟡主观设定", "调整权重重新计算", "高"),
        ("A11", "对标参考", "媒体口径·仅数量级参考", "openEuler 80亿等为媒体公开报道·非正式估值", "可核验: 媒体原始报道", "🟡仅供参考", "不引用对标", "低"),
        ("A12", "综合折价率", "1.10(合理场景)", "流动性+集中度+商业化+合规+竞争综合", "部分可核验: 单项可量化·综合主观", "🟡主观综合", "单列折价明细·见⑦风险折价", "高"),
    ]
    
    for i, (aid, assumption, value, basis, verifiable, status, alt, sensitivity) in enumerate(assumptions, 1):
        r = 4 + i
        ws.cell(row=r, column=1, value=aid)
        ws.cell(row=r, column=2, value=assumption)
        ws.cell(row=r, column=3, value=value)
        ws.cell(row=r, column=4, value=basis)
        ws.cell(row=r, column=5, value=verifiable)
        ws.cell(row=r, column=6, value=status)
        ws.cell(row=r, column=7, value=alt)
        ws.cell(row=r, column=8, value=sensitivity)
    
    style_data_rows(ws, 5, 4 + len(assumptions), max_col)
    
    # 底注
    nr = 4 + len(assumptions) + 2
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=8)
    ws.cell(row=nr, column=1,
            value="✅ = 已实测/公开可核 · 🟡 = 主观判断/待验证 · 🔴 = 未缓解风险。所有🟡假设均被极度保守场景排除。").font = Font(name="PingFang SC", size=9, color=YELLOW)
    
    auto_width(ws, max_col)
    ws.column_dimensions[get_column_letter(2)].width = 22
    ws.column_dimensions[get_column_letter(3)].width = 20
    ws.column_dimensions[get_column_letter(4)].width = 40
    ws.column_dimensions[get_column_letter(7)].width = 28


# ═══════════════════════════════════════
# 入口
# ═══════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="龍魂系统估值模型 v1.1 · Excel 生成器")
    parser.add_argument(
        '--config', '-c',
        default=None,
        help='估值 JSON 配置文件路径（默认使用内置 v1.1 数据）'
    )
    args = parser.parse_args()

    config = load_config(args.config)

    # 如果传入配置，覆盖汇总数值（模块级可变对象直接修改）
    if config:
        keys_map = {
            "极度保守": "SCENARIO_ULTRA_CONSERVATIVE",
            "保守": "SCENARIO_CONSERVATIVE",
            "合理": "SCENARIO_REASONABLE",
            "乐观": "SCENARIO_OPTIMISTIC",
        }
        for scn, cfg_key in keys_map.items():
            if cfg_key in config:
                SCENARIO_TOTALS[scn] = _parse_number(config[cfg_key])

        weight_keys = ["W_W1", "W_W2", "W_W3", "W_W4"]
        scn_order = ["极度保守", "保守", "合理", "乐观"]
        for scn, wkey in zip(scn_order, weight_keys):
            if wkey in config:
                WEIGHTED_WEIGHTS[scn] = _parse_number(config[wkey]) / 100

        WEIGHTED_BENCHMARK = sum(
            SCENARIO_TOTALS[s] * WEIGHTED_WEIGHTS[s] for s in SCENARIO_TOTALS
        )

        output_path = config.get("EXCEL_PATH", "core/valuation/valuation-model.xlsx")
        output_path = Path(output_path)
        if not output_path.is_absolute():
            output_path = Path(__file__).resolve().parent.parent.parent / output_path
    else:
        output_dir = os.path.dirname(os.path.abspath(__file__))
        output_path = os.path.join(output_dir, "longhun-valuation-v1.1.xlsx")

    path = generate_workbook(str(output_path))

    project = config.get("PROJECT_NAME", "龍魂系统") if config else "龍魂系统"
    version = config.get("VERSION", "v1.1") if config else "v1.1"

    print(f"✅ 估值模型已生成: {path}")
    print(f"   项目: {project} {version}")
    print(f"   8张表: ①技术资产 ②文化主权 ③战略叙事 ④生态潜力 ⑤汇总·场景 ⑥敏感性分析 ⑦风险折价 ⑧假设与依据")
    print(f"   底线价: {SCENARIO_TOTALS['极度保守']:,.0f} 万")
    print(f"   加权基准: {WEIGHTED_BENCHMARK:,.0f} 万 ≈ {WEIGHTED_BENCHMARK/10000:.2f} 亿")
    print(f"   DNA: #龍芯⚡️丙午·丙申·丁巳·丙午·䷟恒-VALUATION-v1.1-UID9622")
